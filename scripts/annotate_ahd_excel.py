#!/usr/bin/env python3
"""Annotate the curated AHD workbook without relying on an open Excel handle."""

from __future__ import annotations

import difflib
import os
import re
import shutil
import subprocess
import tempfile
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


REPO = Path(__file__).resolve().parents[1]
PROJECT_ROOT = REPO.parent
WORKBOOK = PROJECT_ROOT / "literature/_data/AHD_papers_full.xlsx"
ARXIV_XML = Path("/tmp/arxiv-selected.xml")
NS = {"a": "http://www.w3.org/2005/Atom"}
EXCLUDED_ARXIV_IDS = {"2605.07039"}  # PACEvolve++ is outside the AHD scope.

GREEN = "C6EFCE"
YELLOW = "FFF2CC"
GRAY = "E7E6E6"
BLUE = "D9EAF7"
HEADER = "243746"


def norm(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def clean_cell(value: str) -> str:
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", " ", value)


def arxiv_id(value: str | None) -> str | None:
    match = re.search(r"(?:arxiv\.org/(?:abs|pdf)/|arXiv:)([0-9]{4}\.[0-9]{4,5})", value or "", re.I)
    return match.group(1) if match else None


def read_site_papers() -> list[dict[str, str]]:
    papers = []
    for path in sorted((REPO / "content/papers").glob("*/index.md")):
        if path.parent.name == "_template":
            continue
        text = path.read_text(encoding="utf-8")
        frontmatter = text.split("---", 2)[1] if text.startswith("---") else ""
        fields = {}
        for key in ("title", "paper_url"):
            match = re.search(rf"^{key}:\s*[\"']?(.+?)[\"']?\s*$", frontmatter, re.M)
            if match:
                fields[key] = match.group(1).strip("\"'")
        year_match = re.search(r"^year:\s*(\d{4})", frontmatter, re.M)
        papers.append({
            "title": fields.get("title", ""),
            "paper_url": fields.get("paper_url", ""),
            "year": int(year_match.group(1)) if year_match else None,
        })
    return papers


def site_match(title: str, url: str, site_papers: list[dict[str, str]]) -> bool:
    wanted_id = arxiv_id(url)
    wanted_title = norm(title)
    for paper in site_papers:
        if wanted_id and wanted_id == arxiv_id(paper["paper_url"]):
            return True
        candidate = norm(paper["title"])
        if wanted_title and candidate and difflib.SequenceMatcher(None, wanted_title, candidate).ratio() >= 0.90:
            return True
    return False


def recent_records() -> list[dict[str, object]]:
    if not ARXIV_XML.exists():
        return []
    root = ET.parse(ARXIV_XML).getroot()
    selected = {
        "2608.07395", "2608.16733", "2608.15546", "2608.12522",
        "2608.10795", "2608.08189",
    }
    records = []
    for entry in root.findall("a:entry", NS):
        ident = entry.findtext("a:id", "", NS).split("/")[-1].split("v")[0]
        if ident not in selected:
            continue
        title = " ".join(entry.findtext("a:title", "", NS).split())
        abstract = " ".join(entry.findtext("a:summary", "", NS).split())
        authors = "; ".join(a.findtext("a:name", "", NS) for a in entry.findall("a:author", NS))
        published = entry.findtext("a:published", "", NS)[:10]
        records.append({
            "importance_rank": None,
            "curation_tier": "recent_arxiv_unreviewed",
            "curation_reason": "recent arXiv search; manual relevance review pending",
            "title": title,
            "year": int(published[:4]),
            "venue": "arXiv",
            "authors": authors,
            "url": f"https://arxiv.org/abs/{ident}",
            "pdf_url": f"https://arxiv.org/pdf/{ident}",
            "pdf_path": None,
            "tags": "llm;algorithm_discovery;program_evolution;recent",
            "categories": "01_representation_design_object;02_search_credit_assignment;06_agentic_discovery_systems",
            "source": "arxiv_recent_2026-08",
            "relevance_score": 10,
            "abstract": abstract,
        })
    return records


def scan_pdf(pdf_path: str | None) -> tuple[str, str, str]:
    if not pdf_path:
        return "unverified", "no_pdf", ""
    path = PROJECT_ROOT / pdf_path
    if not path.exists():
        return "unverified", "no_pdf", ""
    try:
        result = subprocess.run(["pdftotext", "-layout", str(path), "-"], check=True, capture_output=True, text=True)
    except (OSError, subprocess.CalledProcessError):
        return "unverified", "pdf_extract_failed", ""
    text = result.stdout
    body = re.split(r"\n\s*(?:REFERENCES|References|Bibliography)\s*\n", text, maxsplit=1)[0]
    pattern = re.compile(r"\bEoH\b|Evolution of Heuristics|2401\.02051", re.I)
    body_match = pattern.search(body)
    ref_match = pattern.search(text[len(body):]) if len(body) < len(text) else None
    if body_match:
        window = body[max(0, body_match.start() - 180): body_match.end() + 180].replace("\n", " ")
        context = "direct_method" if re.search(r"baseline|method|heuristic|evolution|compare|benchmark", window, re.I) else "related_work"
        return "yes", context, clean_cell(window[:500])
    if ref_match:
        return "yes", "bibliography_only", "reference list only"
    return "no", "not_found", ""


def main() -> None:
    if not WORKBOOK.exists():
        raise SystemExit(f"Workbook not found: {WORKBOOK}")
    # Load and fully close the source workbook before replacing it atomically.
    wb = load_workbook(WORKBOOK)
    ws = wb["papers_curated"]
    headers = [cell.value for cell in ws[1]]
    extra = ["website_status", "eoh_citation", "eoh_citation_context", "eoh_evidence"]
    for name in extra:
        if name not in headers:
            headers.append(name)
            ws.cell(1, len(headers)).value = name
    col = {name: headers.index(name) + 1 for name in headers}
    site_papers = read_site_papers()

    # Remove candidates that were ruled out by scope review, then append only
    # new arXiv IDs/titles so repeated runs remain idempotent.
    for row_number in range(ws.max_row, 1, -1):
        row_url = str(ws.cell(row_number, col["pdf_url"]).value or "")
        if arxiv_id(row_url) in EXCLUDED_ARXIV_IDS:
            ws.delete_rows(row_number)
    existing_keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        existing_keys.add((arxiv_id(str(row[col.get("pdf_url", 10) - 1] or "")), norm(str(row[col["title"] - 1] or ""))))
    for record in recent_records():
        key = (arxiv_id(str(record["pdf_url"])), norm(str(record["title"])))
        if key in existing_keys or any(key[1] == old[1] for old in existing_keys):
            continue
        ws.append([record.get(header) for header in headers[:15]] + [None] * len(extra))
        existing_keys.add(key)

    # The source curation workbook predates a few papers already published on
    # the website. Add lightweight sync rows so the workbook is a complete index.
    for paper in site_papers:
        key = (arxiv_id(paper["paper_url"]), norm(paper["title"]))
        if key in existing_keys or any(key[1] == old[1] for old in existing_keys):
            continue
        record = {
            "importance_rank": None,
            "curation_tier": "website_sync",
            "curation_reason": "already present in the AHD Papers website",
            "title": paper["title"],
            "year": paper.get("year"),
            "venue": "website sync",
            "authors": None,
            "url": paper["paper_url"],
            "pdf_url": paper["paper_url"],
            "pdf_path": None,
            "tags": "website_sync",
            "categories": "",
            "source": "ahd_papers_website",
            "relevance_score": 10,
            "abstract": None,
        }
        ws.append([record.get(header) for header in headers[:15]] + [None] * len(extra))
        existing_keys.add(key)

    # Recompute annotations for every row, including the newly appended records.
    for row in ws.iter_rows(min_row=2, max_col=len(headers)):
        values = {headers[i]: row[i].value for i in range(len(headers))}
        title = str(values.get("title") or "")
        url = str(values.get("pdf_url") or values.get("url") or "")
        is_site = site_match(title, url, site_papers)
        citation, context, evidence = scan_pdf(values.get("pdf_path"))
        # PACE was checked against its arXiv PDF during the recent-paper pass:
        # EoH occurs in the bibliography, which still counts as a citation.
        if arxiv_id(url) == "2608.07395":
            citation, context, evidence = "yes", "bibliography_only", "Checked arXiv PDF; EoH appears only in references."
        if values.get("curation_tier") == "recent_arxiv_unreviewed":
            status = "recent_arxiv_unreviewed"
        else:
            status = "on_website" if is_site else "candidate"
        row[col["website_status"] - 1].value = status
        row[col["eoh_citation"] - 1].value = citation
        row[col["eoh_citation_context"] - 1].value = context
        row[col["eoh_evidence"] - 1].value = evidence
        fill = PatternFill("solid", fgColor=GREEN if citation == "yes" else YELLOW if citation == "unverified" else GRAY)
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[col["website_status"] - 1].fill = PatternFill("solid", fgColor=BLUE)
        if status == "recent_arxiv_unreviewed":
            row[col["website_status"] - 1].fill = PatternFill("solid", fgColor=YELLOW)

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 34
    widths = {"title": 42, "abstract": 72, "eoh_evidence": 48, "website_status": 22, "eoh_citation_context": 22}
    for name, width in widths.items():
        ws.column_dimensions[chr(64 + col[name]) if col[name] <= 26 else "S"].width = width

    if "Legend" in wb.sheetnames:
        del wb["Legend"]
    legend = wb.create_sheet("Legend", 0)
    legend.append(["AHD Papers workbook annotation legend"])
    legend.append(["Color", "Meaning"])
    for label, color, meaning in [
        ("Green", GREEN, "PDF contains an EoH citation anywhere; context records where it appears."),
        ("Yellow", YELLOW, "Recent arXiv candidate or PDF could not be verified."),
        ("Gray", GRAY, "No EoH match found."),
        ("Blue status cell", BLUE, "Paper is already represented on the AHD Papers website."),
    ]:
        legend.append([label, meaning])
        legend.cell(legend.max_row, 1).fill = PatternFill("solid", fgColor=color)
    legend.append([])
    legend.append(["Notes", "Close and reopen Excel after this script finishes so its cached workbook view cannot overwrite the atomic replacement."])
    legend.column_dimensions["A"].width = 22
    legend.column_dimensions["B"].width = 110
    for row in legend.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    legend.freeze_panes = "A3"

    fd, temp_name = tempfile.mkstemp(prefix="AHD_papers_full.", suffix=".xlsx", dir=str(WORKBOOK.parent))
    os.close(fd)
    try:
        wb.save(temp_name)
        shutil.copy2(WORKBOOK, WORKBOOK.with_suffix(".before-annotation.xlsx"))
        os.replace(temp_name, WORKBOOK)
    finally:
        if os.path.exists(temp_name):
            os.unlink(temp_name)
    print(f"Wrote {WORKBOOK} with {ws.max_row - 1} records and {len(headers)} columns")


if __name__ == "__main__":
    main()
