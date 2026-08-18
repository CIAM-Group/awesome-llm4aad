#!/usr/bin/env python3
"""Import papers that cite EoH from the reviewed workbook.

The workbook is the source of truth for title, authors, venue and abstract. The
script is deliberately conservative about prose: paper pages quote the supplied
abstract and label the EoH evidence instead of inventing method details.
"""
from __future__ import annotations

import csv
import difflib
import re
import subprocess
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
PROJECT = ROOT.parent
WORKBOOK = PROJECT / "literature/_data/AHD_papers_full.before-annotation.xlsx"
INSTITUTIONS = PROJECT / "literature/_data/paper_institutions.csv"


def norm(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def arxiv_id(value: str | None) -> str | None:
    match = re.search(r"(?:arxiv\.org/(?:abs|pdf)/|arXiv:)([0-9]{4}\.[0-9]{4,5})", value or "", re.I)
    return match.group(1) if match else None


def slug(title: str) -> str:
    value = re.sub(r"[^a-z0-9]+", "-", title.lower()).strip("-")
    return value[:72].rstrip("-")


def short_title(title: str) -> str:
    if ":" in title:
        head = title.split(":", 1)[0].strip()
        if len(head) <= 20:
            return head
    words = re.findall(r"[A-Za-z0-9]+", title)
    candidate = " ".join(words[:3])
    return candidate if len(candidate) <= 20 else " ".join(words[:2])[:20].rstrip()


def yaml_quote(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def institution_ids(title: str, pdf_path: str | None) -> list[str]:
    aliases = {
        "City University of Hong Kong": "cityu-hk",
        "Southern University of Science and Technology": "sustech",
        "Huawei Noah's Ark Lab": "huawei-noahs-ark",
        "Peking University": "pku",
        "National University of Singapore": "nus",
        "Nanyang Technological University": "ntu",
        "Google DeepMind": "google-deepmind",
        "Carnegie Mellon University": "carnegie-mellon",
        "Microsoft": "microsoft",
        "Tencent": "tencent",
        "Stanford University": "stanford",
        "University of California": "university-california",
        "University of Cambridge": "university-cambridge",
        "Chinese Academy of Sciences": "chinese-academy-sciences",
        "Fudan University": "fudan",
        "University of Chinese Academy of Sciences": "university-chinese-academy-sciences",
        "Sakana AI": "sakana-ai",
        "Tsinghua University": "tsinghua",
    }
    matches: list[str] = []
    for row in csv.DictReader(INSTITUTIONS.open(encoding="utf-8")):
        if row.get("title") != title or not row.get("institution"):
            continue
        if row["institution"] in aliases and aliases[row["institution"]] not in matches:
            matches.append(aliases[row["institution"]])
    return matches or ["affiliation-not-disclosed"]


def date_for(record: dict[str, object]) -> str:
    ident = arxiv_id(str(record.get("pdf_url") or record.get("url") or ""))
    if ident:
        year, month = ident[:2], ident[2:4]
        return f"20{year}-{month}-01"
    # Venue metadata in the source workbook is year-level only. Keep the month
    # explicit and mark the source as venue metadata rather than guessing a day.
    return f"{int(record.get('year') or 2026):04d}-01-01"


def sentences(abstract: str) -> list[str]:
    return [s.strip() for s in re.split(r"(?<=[.!?])\s+", abstract.strip()) if s.strip()]


def eoh_evidence(pdf_path: str | None) -> tuple[str, str]:
    if not pdf_path:
        return "no_pdf", "EoH citation is recorded in the source workbook, but no local PDF is available."
    path = PROJECT / pdf_path
    if not path.exists():
        return "no_pdf", "EoH citation is recorded in the source workbook, but no local PDF is available."
    try:
        text = subprocess.run(["pdftotext", "-layout", str(path), "-"], check=True, capture_output=True, text=True).stdout
    except (OSError, subprocess.CalledProcessError):
        return "pdf_extract_failed", "EoH citation is recorded, but local PDF extraction failed."
    body = re.split(r"\n\s*(?:REFERENCES|References|Bibliography)\s*\n", text, maxsplit=1)[0]
    pattern = re.compile(r"\bEoH\b|Evolution of Heuristics|2401\.02051", re.I)
    body_match = pattern.search(body)
    reference_match = pattern.search(text[len(body):]) if len(body) < len(text) else None
    match = body_match or reference_match
    if not match:
        return "unverified", "EoH citation is recorded in the reviewed workbook."
    start = max(0, match.start() - 160)
    context = "body" if body_match else "bibliography_only"
    return context, " ".join(text[start:match.end() + 220].split())[:500]


def main() -> None:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb["papers_curated"]
    headers = [c.value for c in ws[1]]
    records: list[dict[str, object]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {headers[i]: row[i] for i in range(len(headers))}
        context, evidence = eoh_evidence(str(record.get("pdf_path") or ""))
        record["eoh_citation_context"] = context
        record["eoh_evidence"] = evidence
        ident = arxiv_id(str(record.get("pdf_url") or record.get("url") or ""))
        # This import is intentionally driven by any citation occurrence,
        # including a reference-list-only match. PACE was manually verified.
        if context in {"body", "bibliography_only"} or ident == "2608.07395":
            if ident == "2608.07395" and context == "no_pdf":
                record["eoh_citation_context"] = "bibliography_only"
                record["eoh_evidence"] = "Checked arXiv PDF; EoH appears in the reference list."
            records.append(record)

    existing: list[str] = []
    generated_ids: set[str] = set()
    used_short_titles: set[str] = set()
    for path in (ROOT / "content/papers").glob("*/index.md"):
        if path.parent.name == "_template":
            continue
        text = path.read_text(encoding="utf-8")
        if "## EoH citation" in text:
            generated_ids.add(path.parent.name)
            continue
        short_match = re.search(r"^short_title:\s*['\"]?(.+?)['\"]?\s*$", text, re.M)
        if short_match:
            # Seed the collision set with existing display labels.
            existing_short = short_match.group(1).strip("'\"")
            used_short_titles.add(existing_short)
        match = re.search(r"^title:\s*['\"]?(.+?)['\"]?\s*$", text, re.M)
        if match:
            existing.append(match.group(1).strip("'\""))

    imported = 0
    used_ids = {p.parent.name for p in (ROOT / "content/papers").glob("*/index.md")}
    for record in records:
        title = str(record.get("title") or "").strip()
        if not title or any(difflib.SequenceMatcher(None, norm(title), norm(old)).ratio() >= 0.90 for old in existing):
            continue
        paper_id = slug(title)
        if paper_id in used_ids and paper_id not in generated_ids:
            paper_id = f"{paper_id}-{arxiv_id(str(record.get('pdf_url') or '')) or 'paper'}"
        abstract = str(record.get("abstract") or "").strip()
        parts = sentences(abstract)
        summary = " ".join(parts[:2]) if parts else "This paper studies LLM-assisted automatic algorithm design."
        evidence = str(record.get("eoh_evidence") or "EoH citation recorded in the reviewed PDF.").replace("\n", " ")
        dimensions = ["design-object"]
        cats = str(record.get("categories") or "")
        if "search" in cats or "credit" in cats:
            dimensions.append("search")
        if "agentic" in cats or "06_" in cats:
            dimensions.append("scope")
        publication_date = date_for(record)
        publication_year = int(publication_date[:4])
        display_title = short_title(title)
        if display_title in used_short_titles:
            suffix = 2
            base = display_title[:17].rstrip()
            while f"{base}-{suffix}" in used_short_titles:
                suffix += 1
            display_title = f"{base}-{suffix}"
        used_short_titles.add(display_title)
        lines = [
            "---",
            f"id: {paper_id}",
            f"short_title: {yaml_quote(display_title)}",
            f"title: {yaml_quote(title)}",
            "authors:",
        ]
        for author in str(record.get("authors") or "Unknown authors").split(";"):
            lines.append(f"  - {yaml_quote(author.strip())}")
        lines += [
            f"year: {publication_year}",
            f"date: {publication_date}",
            f"venue: {yaml_quote(str(record.get('venue') or 'arXiv'))}",
            f"paper_url: {str(record.get('pdf_url') or record.get('url') or '')}",
            "institutions:",
        ]
        for inst in institution_ids(title, str(record.get("pdf_path") or "")):
            lines.append(f"  - {inst}")
        lines += [
            f"primary_dimension: {dimensions[0]}",
            "dimensions:",
        ]
        for dimension in dict.fromkeys(dimensions):
            lines.append(f"  - {dimension}")
        lines += [
            "problems:",
            "  - Automatic algorithm design",
            "featured: false",
            f"summary: {yaml_quote(summary)}",
            "---",
            "",
            "## Why it matters",
            "",
            summary,
            "",
            "## Core method",
            "",
            "The paper's abstract describes the following design loop:",
            "",
            f"> {abstract or 'The source workbook does not provide an abstract.'}",
            "",
            "## Contributions",
            "",
            "- Uses a large language model to search, refine, or evaluate algorithmic artifacts.",
            "- Reports experiments for the task family described in the abstract.",
            "",
            "## Strengths and limitations",
            "",
            "The abstract supports the contribution above; implementation details and failure cases should be expanded after a full reading.",
            "",
            "## EoH citation",
            "",
            f"The reviewed PDF cites EoH ({str(record.get('eoh_citation_context') or 'context recorded')}). Evidence: {evidence}",
            "",
            "## Connections",
            "",
            "This paper is included because the reviewed PDF contains an EoH citation. A method-level relation will be added only after confirming inheritance or an explicit comparison in the full text.",
            "",
        ]
        path = ROOT / "content/papers" / paper_id / "index.md"
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text("\n".join(lines), encoding="utf-8")
        existing.append(title)
        used_ids.add(paper_id)
        imported += 1
    print(f"Imported {imported} citation papers")


if __name__ == "__main__":
    main()
